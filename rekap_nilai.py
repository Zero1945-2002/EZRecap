import PySimpleGUI as sg
from openpyxl import load_workbook
import os
import subprocess
import win32com.client



TR = ['TR1', 'TR2', 'TR3', 'TR4', 'TR5', 'TR6', 'TR7', 'TR8', 'TR9', 'TR10']
TA = ['TA1', 'TA2', 'TA3', 'TA4', 'TA5', 'TA6', 'TA7', 'TA8', 'TA9', 'TA10']

file_types = [("Excel Workbook (*.xlsx)", "*.xlsx"),
              ("Excel 97-2003 Workbook (*.xls)", "*.xls"),
              ("All files (*.*)", "*.*")]

halaman_awal = [[
            sg.Text("Excel File"),
            sg.Input(tooltip='klik tombol browse untuk memilih file',default_text='',size=(25, 1), key="-FILE-"),
            sg.FileBrowse(file_types=file_types),
            sg.Button("Buka Excel")
        ]]

window = sg.Window("Rekap Nilai", halaman_awal)

def is_file_open(filename):
    try:
        # Mencoba membuka file dalam mode penulisan (write mode)
        with open(filename, 'a') as f:
            pass
        return False  # Jika berhasil membuka file, maka file tidak dibuka oleh aplikasi lain
    except PermissionError:
        return True   # Jika mendapatkan exception PermissionError, maka file dibuka oleh aplikasi lain


def update_nilai_tr():
    if len(values['-KELAS-']) == 0:
        sg.popup('Mohon pilih kelasnya terlebih dahulu!')
        return
    name = str(values['-KELAS-'][0])
    ws = workbook[name]
    for k in ws.iter_cols():
        for c in k:
            if c.value == "NIM" or c.value == "nim":
                nim_col = c.column
    for row in ws.iter_rows(min_col=nim_col, max_col=nim_col):
        for cell in row:
            if str(values['-NIM-']) == ' ' or str(values['-NIM-']) == '':
                sg.popup('Masukkan NIM terlebih dahulu!')
                return
            elif str(cell.value) == str(values['-NIM-']):
                for kolom in ws.iter_cols():
                    for cel in kolom:
                        if len(values['-TR-']) == 0:
                            sg.popup('Mohon pilih TR terlebih dahulu!')
                            return
                        if cel.value == values['-TR-'][0]:
                            if str(values['-NILAI-TR-']).isdigit():
                                nilai = int(values['-NILAI-TR-'])
                                if nilai < 0 or nilai > 100:
                                    sg.popup('Masukkan nilai dengan range 0 s/d 100')
                                    return
                            else:
                                sg.popup('Masukkan nilai dengan range 0 s/d 100')
                                return
                            ws.cell(row=cell.row, column=cel.column, value=int(values['-NILAI-TR-']))
                            workbook.save(filename=filename)
                            for m in ws.iter_cols():
                                for n in m:
                                    if n.value == "Nama" or n.value == "nama" or n.value == "NAMA":
                                        nama_col = n.column
                            nama_praktikan = ws.cell(row=cell.row, column=nama_col).value
                            nim_praktikan = cell.value
                            teks = 'Nilai '+str(values['-TR-'][0])+' praktikan '+str(nama_praktikan)+' ('+str(nim_praktikan)+') telah di update!'
                            sg.popup(teks)
                            return
                sg.popup('Pastikan kolom '+str(values['-TR-'][0])+' tersedia!')
                return
    sg.popup('Praktikan dengan NIM '+str(values['-NIM-'])+' tidak ditemukan!')
    return
            
def update_nilai_ta():
    if len(values['-KELAS-']) == 0:
        sg.popup('Mohon pilih kelasnya terlebih dahulu!')
        return
    name = str(values['-KELAS-'][0])
    ws = workbook[name]
    for k in ws.iter_cols():
        for c in k:
            if c.value == "NIM" or c.value == "nim":
                nim_col = c.column
    for row in ws.iter_rows(min_col=nim_col, max_col=nim_col):
        for cell in row:
            if str(values['-NIM-']) == ' ' or str(values['-NIM-']) == '':
                sg.popup('Masukkan NIM terlebih dahulu!')
                return
            elif str(cell.value) == str(values['-NIM-']):
                for kolom in ws.iter_cols():
                    for cel in kolom:
                        if len(values['-TA-']) == 0:
                            sg.popup('Mohon pilih TA terlebih dahulu!')
                            return
                        if cel.value == values['-TA-'][0]:
                            if str(values['-NILAI-TA-']).isdigit():
                                nilai = int(values['-NILAI-TA-'])
                                if nilai < 0 or nilai > 100:
                                    sg.popup('Masukkan nilai dengan range 0 s/d 100')
                                    return
                            else:
                                sg.popup('Masukkan nilai dengan range 0 s/d 100')
                                return
                            ws.cell(row=cell.row, column=cel.column, value=int(values['-NILAI-TA-']))
                            workbook.save(filename=filename)
                            for m in ws.iter_cols():
                                for n in m:
                                    if n.value == "Nama" or n.value == "nama" or n.value == "NAMA":
                                        nama_col = n.column
                            nama_praktikan = ws.cell(row=cell.row, column=nama_col).value
                            nim_praktikan = ws.cell(row=cell.row, column=3).value
                            teks = 'Nilai '+str(values['-TA-'][0])+' praktikan '+str(nama_praktikan)+' ('+str(nim_praktikan)+') telah di update!'
                            sg.popup(teks)
                            return
                sg.popup('Pastikan kolom '+str(values['-TA-'][0])+' tersedia!')
                return
    sg.popup('Praktikan dengan NIM '+str(values['-NIM-'])+' tidak ditemukan!')
    return
            
while True:
    event, values = window.read()
    print(event, values)

    if event == sg.WIN_CLOSED:
        break
    elif event == "Buka Excel":
        if values['-FILE-'] == ' ' or values['-FILE-'] == '':
            sg.popup('Mohon pilih filenya terlebih dahulu!')
        else:
            filename = values["-FILE-"]
            if os.path.isfile(filename):
                nama_file = filename
                if is_file_open(nama_file):
                    excel_app = win32com.client.Dispatch("Excel.Application")
                    wb = excel_app.Workbooks.Open(nama_file)
                    wb.Save()
                    wb.Close(SaveChanges=False)
                    excel_app.Quit()
                workbook = load_workbook(filename=nama_file)
                sheet = workbook.sheetnames
                col1 = sg.Column([
                                [sg.Frame('Kelas:',[[sg.Text('Kelas:'),sg.Listbox(sheet, size=(20, 2), enable_events=True, key='-KELAS-')],],size=(505,70))],
                                [sg.Frame('NIM:',[[sg.Text('NIM:'),sg.Input(tooltip='ex : 202111129', default_text='', key='-NIM-', size=(19,1))]],size=(505,50))]], size=(515,140), pad=(0,0))                                  
                col2 = sg.Column([[sg.Frame('Tugas Rumah:', [[sg.Text(), sg.Column([
                                                                                    [sg.Listbox(TR, size=(20, 5), enable_events=True, key='-TR-')],
                                                                                    [sg.Text('Nilai:'),sg.Input(tooltip='masukkan nilai 1 s/d 100', default_text='', key='-NILAI-TR-', size=(19,1))],
                                                                                    [sg.Button('Simpan Nilai TR')]], size=(230,155), pad=(0,0))]])] ], pad=(0,0))
                col3 = sg.Column([[sg.Frame('Test Awal:', [[sg.Text(), sg.Column([
                                                                                [sg.Listbox(TA, size=(20, 5), enable_events=True, key='-TA-')],
                                                                                [sg.Text('Nilai:'),sg.Input(tooltip='masukkan nilai 1 s/d 100', default_text='', key='-NILAI-TA-', size=(19,1))],
                                                                                [sg.Button('Simpan Nilai TA')]], size=(230,155), pad=(0,0))]])] ], pad=(0,0))
                col4 = sg.Column([[sg.Frame('Actions:',
                                                    [[sg.Column([[sg.Button('Kembali'), sg.Button('Buka File')]],size=(460,45), pad=(0,0))]],size=(505,60))]], pad=(0,0))
                halaman_lanjut = [[col1],
                        [col2, col3],
                        [col4]]
                
                window.close()
                window = sg.Window('Rekap Nilai', halaman_lanjut)
            else:
                sg.popup('File tidak ditemukan!')
    elif event == 'Simpan Nilai TR':
        if is_file_open(nama_file):
            excel_app = win32com.client.Dispatch("Excel.Application")
            wb = excel_app.Workbooks.Open(nama_file)
            wb.Close(SaveChanges=True)
            excel_app.Quit()
        update_nilai_tr()
    elif event == "Simpan Nilai TA":
        if is_file_open(nama_file):
            excel_app = win32com.client.Dispatch("Excel.Application")
            wb = excel_app.Workbooks.Open(nama_file)
            wb.Close(SaveChanges=True)
            excel_app.Quit()
        update_nilai_ta()
    elif event == 'Kembali':
        halaman_awal = [[
            sg.Text("Excel File"),
            sg.Input(tooltip='klik tombol browse untuk memilih file',default_text='',size=(25, 1), key="-FILE-"),
            sg.FileBrowse(file_types=file_types),
            sg.Button("Buka Excel")
        ]]
        window.close()
        window = sg.Window('Rekap Nilai', halaman_awal)
    elif event == 'Buka File':
        try:
            # Membuka file Excel dengan aplikasi yang terkait
            subprocess.Popen(['start', '', nama_file], shell=True)
            sg.popup('File berhasil dibuka!')
            
        except FileNotFoundError:
            print(f"File tidak ditemukan")
        except Exception as e:
            print("Terjadi kesalahan:", str(e))


window.close()