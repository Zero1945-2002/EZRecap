import sys
import os
import subprocess
from asset.GUI.gui import Ui_MainWindow
from openpyxl import load_workbook
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QFileDialog, QTextEdit, QMessageBox
from PyQt6.QtGui import QIcon, QPixmap

file_types = [
    ("Excel Workbook (*.xlsx)", "*.xlsx"),  # Format file Excel xlsx
    ("Excel 97-2003 Workbook (*.xls)", "*.xls"),  # Format file Excel xls (versi lama)
    ("Excel Macro-Enabled Workbook (*.xlsm)", "*.xlsm"),  # Format file Excel xlsm (mengandung macro)
    ("Excel Template (*.xltx)", "*.xltx"),  # Format file Excel xltx (template)
    ("Excel Macro-Enabled Template (*.xltm)", "*.xltm"),  # Format file Excel xltm (template mengandung macro)
    ("CSV (Comma Delimited) (*.csv)", "*.csv"),  # Format file Excel csv (Comma Separated Values)
    ("Text File (*.txt)", "*.txt"),  # Format file Excel txt (Text File)
    ("PRN (Printable) (*.prn)", "*.prn"),  # Format file Excel prn (Printable)
    ("All files (*.*)", "*.*")  # Semua jenis file
]


class FileBrowser(QWidget):
    def __init__(self):
        super().__init__()
        filenames = QFileDialog.getOpenFileName(self, "Choose File", "", ";;".join([f"{desc} ({ext})" for desc, ext in file_types]))
        self.filename = filenames

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon(QPixmap("asset/icons/favico.ico")))
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        ui = self.ui

        ui.Kelas.setEnabled(False)
        ui.NIM.setEnabled(False)
        ui.TR.setEnabled(False)
        ui.TA.setEnabled(False)
        ui.Nilai_TR.setEnabled(False)
        ui.Nilai_TA.setEnabled(False)
        ui.Simpan_TR.setEnabled(False)
        ui.Simpan_TA.setEnabled(False)
        ui.Buka_File.setEnabled(False)

        ui.Kelas.addItem("-None-")
        ui.TR.addItem("-None-")
        ui.TA.addItem("-None-")

        ui.Open_File.triggered.connect(self.open_file)
        ui.Kelas.currentTextChanged.connect(self.pilihan_kelas)
        ui.NIM.textChanged.connect(lambda: self.validate_input(ui.NIM))
        ui.Nilai_TR.textChanged.connect(lambda: self.validate_input(ui.Nilai_TR))
        ui.Nilai_TA.textChanged.connect(lambda: self.validate_input(ui.Nilai_TA))
        ui.TR.currentTextChanged.connect(self.open_tr)
        ui.TA.currentTextChanged.connect(self.open_ta)
        ui.Simpan_TR.clicked.connect(lambda: self.simpan_nilai(ui.Simpan_TR))
        ui.Simpan_TA.clicked.connect(lambda: self.simpan_nilai(ui.Simpan_TA))
        ui.Buka_File.clicked.connect(self.buka_file)
        ui.About.triggered.connect(self.about)

    def about(self):
        self.warning_message(['EZRecap',
                 'Alat bantu asisten lab dalam menginput nilai praktikan',
                 'dibuat oleh Zaid Immaduddin Abdurrahman',
                 '',
                 'contact us :',
                 'Email : zaidimmaduddin56@gmail.com',
                 'GitHub : https://github.com/zaid-2002/EZRecap.git'])

    def buka_file(self):
        global filename
        try:
            # Membuka file Excel dengan aplikasi yang terkait
            result = subprocess.run(['start', '', filename], shell=True, capture_output=True, text=True)
            if result.stderr:
                error_message = result.stderr.strip()
                if "The process cannot access the file because it is being used by another process." in error_message:
                    self.warning_message('File sedang digunakan oleh proses lain dan telah terbuka!')
                else:
                    self.warning_message('Error:', error_message)
            else:
                self.warning_message('File berhasil dibuka!')
        except PermissionError as e:
                self.warning_message('PermissionError:', str(e))
        except FileNotFoundError:
            self.warning_message(f"File tidak ditemukan")
        except Exception as e:
            self.warning_message("Mohon pilih filenya terlebih dahulu!")
        pass

    def simpan_nilai(self, widget:QWidget):
        if widget == self.ui.Simpan_TR:
            t_type = self.ui.TR
            n_type = self.ui.Nilai_TR
        elif widget == self.ui.Simpan_TA:
            t_type = self.ui.TA
            n_type = self.ui.Nilai_TA
        nama = False
        nim = False
        p_nim = False
        jenis = False

        # NIM
        if str(self.ui.NIM.toPlainText().strip()) == '':
            self.warning_message('Masukkan NIM terlebih dahulu!')
            return

        # Nilai
        nilai = int(n_type.toPlainText())
        if nilai < 0 or nilai > 100:
            self.warning_message('Masukkan nilai dengan range 0 s/d 100')
            return

        # Pencarian cell
        kelas = str(self.ui.Kelas.currentText())
        ws = workbook[kelas]
        for k in ws.iter_cols():
            for c in k:
                if c.value == "NIM" or c.value == "nim" or c.value == "Nim":
                    nim_col = c.column
                    nim = True
                    for row in ws.iter_rows(min_col=nim_col, max_col=nim_col):
                        for cell in row:
                            if str(cell.value) == str(self.ui.NIM.toPlainText()):
                                nim_row = cell.row
                                p_nim = True
                if c.value == t_type.currentText():
                    t_col = c.column
                    jenis = True
                if c.value == "Nama" or c.value == "NAMA" or c.value == "nama":
                    nama_col = c.column
                    nama = True
        if self.cek(nim,"Pastikan kolom 'NIM' tersedia!") == False:
            return
        if self.cek(p_nim,'Praktikan dengan nim '+str(self.ui.NIM.toPlainText())+' tidak ditemukan!') == False:
            return
        if self.cek(jenis,"Pastikan kolom '"+str(t_type.currentText())+"' tersedia!") == False:
            return
        if self.cek(nama,"Pastikan kolom 'Nama' tersedia!") == False:
            return

        # Ubah nilai cell
        ws.cell(row=nim_row, column=t_col, value=int(n_type.toPlainText()))
        try:
            workbook.save(filename=filename)
        except PermissionError:
            self.warning_message('Harap tutup terlebih dahulu file yang terbuka!')
            return
        nama_praktikan = ws.cell(row=nim_row, column=nama_col).value
        nim_praktikan = ws.cell(row=nim_row, column=nim_col).value
        teks = 'Nilai '+str(t_type.currentText())+' praktikan '+str(nama_praktikan)+' ('+str(nim_praktikan)+') berhasil dimasukkan!'
        self.warning_message(teks)
        return

    def cek(self,var,text):
        if var == False:
            self.warning_message(text)
            return False
        return True

    def warning_message(self, text):
        msg = QMessageBox()
        if type(text) == list:
            kalimat = "\n".join(text)
            msg.setWindowTitle("About")
            msg.setText(kalimat)
        else:
            msg.setWindowTitle("Notifikasi")
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setText(text)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def validate_input(self, widget:QTextEdit):
        cursor_position = widget.textCursor().position()
        text = widget.toPlainText()
        filtered_text = ''.join(c for c in text if c.isdigit())
        widget.blockSignals(True)
        widget.setPlainText(filtered_text)
        widget.blockSignals(False)

        cursor = widget.textCursor()
        cursor.setPosition(min(cursor_position, len(filtered_text)))
        widget.setTextCursor(cursor)
        if widget == self.ui.Nilai_TR:
            if not self.ui.Nilai_TR.toPlainText().strip():
                self.ui.Simpan_TR.setEnabled(False)
            else:
                self.ui.Simpan_TR.setEnabled(True)
        if widget == self.ui.Nilai_TA:
            if not self.ui.Nilai_TA.toPlainText().strip():
                self.ui.Simpan_TA.setEnabled(False)
            else:
                self.ui.Simpan_TA.setEnabled(True)

    def open_ta(self):
        if self.ui.TA.currentText() != "-None-":
            self.ui.Nilai_TA.setEnabled(True)
        else:
            self.ui.Nilai_TA.clear()
            self.ui.Nilai_TA.setEnabled(False)

    def open_tr(self):
        if self.ui.TR.currentText() != "-None-":
            self.ui.Nilai_TR.setEnabled(True)
        else:
            self.ui.Nilai_TR.clear()
            self.ui.Nilai_TR.setEnabled(False)

    def pilihan_kelas(self):
        if self.ui.Kelas.currentText() != "-None-":
            self.ui.TR.clear()
            self.ui.TA.clear()
            self.ui.TR.addItem("-None-")
            self.ui.TA.addItem("-None-")
            TR = []
            TA = []
            kelas = self.ui.Kelas.currentText()
            selected_sheet = workbook[kelas]
            for k in selected_sheet.iter_cols():
                for c in k:
                    teks = str(c.value)
                    if teks[0].upper() == "T" and teks[1].upper() == "R":
                        TR.append(c.value)
                    if teks[0].upper() == "T" and teks[1].upper() == "A":
                        TA.append(c.value)
            self.ui.TR.addItems(TR)
            self.ui.TA.addItems(TA)
            self.ui.NIM.setEnabled(True)
            self.ui.Nilai_TR.clear()
            self.ui.Nilai_TA.clear()
            self.ui.TR.setEnabled(True)
            self.ui.TA.setEnabled(True)
        else:
            self.ui.TR.clear()
            self.ui.TA.clear()
            self.ui.TR.addItem("-None-")
            self.ui.TA.addItem("-None-")
            self.ui.Nilai_TR.clear()
            self.ui.Nilai_TA.clear()
            self.ui.NIM.setEnabled(False)
            self.ui.TR.setEnabled(False)
            self.ui.TA.setEnabled(False)


    def open_file(self):
        global filename, workbook
        filenames = FileBrowser()
        filename = filenames.filename[0]
        if filename != None:
            workbook = load_workbook(filename=filename)
            self.setWindowTitle((os.path.basename(filename)) +" - EZRecap")
            sheet = workbook.sheetnames
            self.ui.Kelas.addItems(sheet)
            self.ui.Kelas.setEnabled(True)
            self.ui.Buka_File.setEnabled(True)
            
        

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(QPixmap('asset/icons/favico.ico')))
    window = MainWindow()
    window.show()
    sys.exit(app.exec())